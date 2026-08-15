# /// script
# requires-python = ">=3.11"
# dependencies = [
#   "openpyxl>=3.1",
#   "requests>=2.32",
# ]
# ///

"""DEPRECATED: importer for the retired GardenSuite Brevo workflow.

Do not use this script for current outreach. Read
marketing/outreach/CURRENT_STRATEGY.md.

The script is a dry run unless both --execute and the confirmation phrase are
provided. It reads only the "Email Review" sheet from the reviewed workbook.
"""

from __future__ import annotations

import argparse
import os
import re
import sys
import time
from datetime import date
from pathlib import Path

import requests
from openpyxl import load_workbook


API_URL = "https://api.brevo.com/v3/contacts"
CONFIRMATION = "I_HAVE_REVIEWED_PERMISSION"
DEFAULT_FILE = Path(__file__).resolve().parents[4] / "outputs" / "gardensuite_outreach_ready_20260802" / "GardenSuite_Outreach_Ready_20260802.xlsx"


def text(value: object) -> str:
    return "" if value is None else str(value).strip()


def valid_email(value: str) -> bool:
    return bool(re.fullmatch(r"[^\s@]+@[^\s@]+\.[^\s@]+", value))


def approved_contacts(path: Path) -> tuple[list[dict[str, str]], list[str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "Email Review" not in workbook.sheetnames:
        raise ValueError('Workbook must contain an "Email Review" sheet.')

    rows = workbook["Email Review"].iter_rows(values_only=True)
    headers = [text(value) for value in next(rows)]
    required = {"Estate Name", "Email", "Consent Evidence", "Owner Approval", "Brevo Action", "Sources"}
    missing = required.difference(headers)
    if missing:
        raise ValueError(f"Email Review is missing columns: {', '.join(sorted(missing))}")

    contacts: list[dict[str, str]] = []
    skipped: list[str] = []
    for values in rows:
        row = {headers[index]: text(value) for index, value in enumerate(values)}
        email = row["Email"].lower()
        estate = row["Estate Name"] or email
        if row["Owner Approval"].lower() != "approved":
            continue
        if row["Brevo Action"].lower() not in {"add", "import", "subscribe"}:
            skipped.append(f"{estate}: Brevo Action must be Add, Import, or Subscribe")
            continue
        consent = row["Consent Evidence"]
        if not consent or consent.lower() in {"not recorded", "none", "unknown", "no"}:
            skipped.append(f"{estate}: explicit email consent is not recorded")
            continue
        if not valid_email(email):
            skipped.append(f"{estate}: invalid email")
            continue
        contacts.append(
            {
                "email": email,
                "garden": estate,
                "consent": consent,
                "source": row["Sources"][:200],
            }
        )
    workbook.close()
    return contacts, skipped


def upload(contact: dict[str, str], api_key: str, list_id: int) -> tuple[bool, str]:
    payload = {
        "email": contact["email"],
        "attributes": {
            "GARDEN": contact["garden"],
            "CAMPAIGN": "attendance_nurture",
            "EMAIL_CONSENT": True,
            "CONSENT_DATE": date.today().isoformat(),
            "CONSENT_SOURCE": contact["source"],
            "LEGAL_BASIS": contact["consent"],
            "GARDENSUITE_TAGS": "gardensuite,attendance-page,reviewed-import",
        },
        "listIds": [list_id],
        "updateEnabled": True,
    }
    response = requests.post(
        API_URL,
        headers={"accept": "application/json", "api-key": api_key, "content-type": "application/json"},
        json=payload,
        timeout=30,
    )
    if response.status_code in {201, 204}:
        return True, "created or updated"
    return False, f"HTTP {response.status_code}: {response.text[:300]}"


def arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Dry-run or import approved GardenSuite email opt-ins.")
    parser.add_argument("--file", type=Path, default=DEFAULT_FILE, help="Reviewed outreach workbook.")
    parser.add_argument("--execute", action="store_true", help="Perform the Brevo API writes.")
    parser.add_argument("--confirm", default="", help=f"Required with --execute: {CONFIRMATION}")
    return parser.parse_args()


def main() -> int:
    args = arguments()
    if not args.file.is_file():
        print(f"ERROR: Workbook not found: {args.file}")
        return 2

    contacts, skipped = approved_contacts(args.file)
    print(f"Workbook: {args.file}")
    print(f"Approved contacts eligible for automated email: {len(contacts)}")
    print(f"Approved rows rejected by safety checks: {len(skipped)}")
    for reason in skipped[:10]:
        print(f"  SKIP: {reason}")

    if not args.execute:
        print("DRY RUN ONLY. No Brevo contact was changed.")
        return 0
    if args.confirm != CONFIRMATION:
        print(f"ERROR: --execute also requires --confirm {CONFIRMATION}")
        return 2
    if not contacts:
        print("Nothing to import.")
        return 0

    api_key = os.getenv("BREVO_API_KEY", "").strip()
    list_value = os.getenv("BREVO_LIST_ID", "").strip()
    if not api_key or not list_value.isdigit():
        print("ERROR: BREVO_API_KEY and numeric BREVO_LIST_ID must be set.")
        return 2

    successes = 0
    for contact in contacts:
        ok, message = upload(contact, api_key, int(list_value))
        print(f"{'OK' if ok else 'ERROR'} {contact['email']}: {message}")
        successes += int(ok)
        time.sleep(0.25)
    print(f"Imported {successes} of {len(contacts)} approved contacts.")
    return 0 if successes == len(contacts) else 1


if __name__ == "__main__":
    sys.exit(main())
