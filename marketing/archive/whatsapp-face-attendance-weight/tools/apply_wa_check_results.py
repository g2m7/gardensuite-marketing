"""DEPRECATED: result applier for a retired campaign tracker.

Do not use for current outreach. Read marketing/outreach/CURRENT_STRATEGY.md.
"""

import argparse
import csv
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROOT = Path(__file__).resolve().parents[4]
CAMPAIGN_DIR = ROOT / "marketing" / "whatsapp-campaigns" / "face-attendance-weight"
WORKBOOK = CAMPAIGN_DIR / "GS_CRM_WhatsApp_Campaign_Tracker.xlsx"
BACKUP_DIR = CAMPAIGN_DIR / f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}_pre_apply_wa_results"
GRID = "D7E5DC"
TEXT = "111111"


def normalize_phone(value):
    if value is None:
        return ""
    digits = re.sub(r"\D", "", str(value))
    if len(digits) == 12 and digits.startswith("91"):
        digits = digits[2:]
    if len(digits) == 11 and digits.startswith("0"):
        digits = digits[1:]
    if len(digits) == 10 and digits[0] in "6789":
        return digits
    return ""


def get_headers(ws):
    return {str(ws.cell(1, col).value).strip(): col for col in range(1, ws.max_column + 1) if ws.cell(1, col).value}


def style_cell(cell):
    cell.font = Font(name="Aptos", size=10, color=TEXT)
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    cell.border = Border(bottom=Side(style="thin", color=GRID))


def parse_dt(value):
    if not value:
        return datetime.now()
    try:
        return datetime.fromisoformat(value)
    except Exception:
        return datetime.now()


def main():
    parser = argparse.ArgumentParser(description="Apply reviewed WhatsApp check results to tracker workbook.")
    parser.add_argument("results_csv", type=Path)
    parser.add_argument("--apply-no-whatsapp-hold", action="store_true", default=True)
    args = parser.parse_args()

    if not args.results_csv.exists():
        raise SystemExit(f"Results CSV not found: {args.results_csv}")

    BACKUP_DIR.mkdir(parents=True, exist_ok=False)
    backup = BACKUP_DIR / WORKBOOK.name
    backup.write_bytes(WORKBOOK.read_bytes())

    wb = load_workbook(WORKBOOK)
    ws = wb["WhatsApp Tracker"]
    h = get_headers(ws)
    required = [
        "WhatsApp Number",
        "WA Check Status",
        "WA Checked At",
        "WA Check Notes",
        "Send Decision",
        "Priority",
        "Final Outcome",
        "Next Action",
        "Sort Score",
    ]
    missing = [name for name in required if name not in h]
    if missing:
        raise SystemExit(f"Workbook is missing columns: {', '.join(missing)}")

    phone_to_row = {}
    for row in range(2, ws.max_row + 1):
        phone = normalize_phone(ws.cell(row, h["WhatsApp Number"]).value)
        if phone and phone not in phone_to_row:
            phone_to_row[phone] = row

    applied = 0
    no_whatsapp = 0
    with args.results_csv.open("r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for result in reader:
            status = (result.get("wa_check_status") or "").strip()
            if status in {"", "Dry run"}:
                continue
            phone = normalize_phone(result.get("phone"))
            row = phone_to_row.get(phone)
            if not row:
                continue

            ws.cell(row, h["WA Check Status"], status)
            ws.cell(row, h["WA Checked At"], parse_dt(result.get("wa_checked_at")))
            ws.cell(row, h["WA Check Notes"], result.get("wa_check_notes") or "")
            ws.cell(row, h["WA Checked At"]).number_format = "yyyy-mm-dd hh:mm"
            style_cell(ws.cell(row, h["WA Check Status"]))
            style_cell(ws.cell(row, h["WA Checked At"]))
            style_cell(ws.cell(row, h["WA Check Notes"]))

            if status == "No WhatsApp" and args.apply_no_whatsapp_hold:
                ws.cell(row, h["Send Decision"], "Hold")
                ws.cell(row, h["Priority"], "Hold")
                ws.cell(row, h["Final Outcome"], "Disqualified")
                ws.cell(row, h["Next Action"], "Do not send")
                # Worst practical source/action sort bucket.
                ws.cell(row, h["Sort Score"], 9999)
                no_whatsapp += 1

            applied += 1

    wb.save(WORKBOOK)
    print(f"Applied rows: {applied}")
    print(f"No WhatsApp moved to Hold: {no_whatsapp}")
    print(f"Backup: {backup}")
    print(f"Workbook: {WORKBOOK}")


if __name__ == "__main__":
    main()
