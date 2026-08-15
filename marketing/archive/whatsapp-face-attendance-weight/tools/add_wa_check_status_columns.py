from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PATH = Path(r"C:\projects\biz\gardensuite.in\marketing\whatsapp-campaigns\face-attendance-weight\GS_CRM_WhatsApp_Campaign_Tracker.xlsx")
HEADER = "073B28"
GRID = "D7E5DC"
TEXT = "111111"


wb = load_workbook(PATH)
ws = wb["WhatsApp Tracker"]
existing = {str(ws.cell(1, c).value).strip(): c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}

needed = ["WA Check Status", "WA Checked At", "WA Check Notes"]
for header in needed:
    if header not in existing:
        col = ws.max_column + 1
        ws.cell(1, col, header)
        existing[header] = col

for header in needed:
    col = existing[header]
    cell = ws.cell(1, col)
    cell.fill = PatternFill("solid", fgColor=HEADER)
    cell.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(bottom=Side(style="thin", color="9DB7A7"))
    for row in range(2, ws.max_row + 1):
        c = ws.cell(row, col)
        if header == "WA Check Status" and c.value in (None, ""):
            c.value = "Unchecked"
        c.font = Font(name="Aptos", size=10, color=TEXT)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = Border(bottom=Side(style="thin", color=GRID))
        if header == "WA Checked At":
            c.number_format = "yyyy-mm-dd hh:mm"

widths = {
    "WA Check Status": 16,
    "WA Checked At": 18,
    "WA Check Notes": 34,
}
for header, width in widths.items():
    ws.column_dimensions[ws.cell(1, existing[header]).column_letter].width = width

ws.auto_filter.ref = f"A1:{ws.cell(1, ws.max_column).column_letter}{ws.max_row}"
wb.save(PATH)
print("WA check columns:", {h: ws.cell(1, existing[h]).column_letter for h in needed})
