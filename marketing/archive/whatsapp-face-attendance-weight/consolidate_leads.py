"""
GardenSuite Lead Consolidation Script v2
=========================================
Merges all lead sources into one clean Excel file.

Key improvement over v1:
  - Uses union-find clustering so the same estate appearing with different
    phone numbers or emails across sources is collapsed into ONE row.
  - Keeps up to 3 phone numbers and 3 email addresses per estate.
  - All other fields (address, location, CRM status, WA status) are merged
    from all matching rows - best available data wins.

Sources:
  1. tea_estate_contacts_v2.xlsx       (best: phone + email + address)
  2. tea_estate_contacts.xlsx          (phone + address)
  3. Tea Estates.xlsx                  (names + area, Dibrugarh govt list)
  4. Tea Estates number required.xlsx  (names + phone)
  5. email assam.dooars teaestate.xlsx (emails only)
  6. verified_emails_v2.csv            (verified emails + estate match)
  7. verified_list.csv                 (high-confidence verified emails)
  8. midsize_verified.csv              (verified midsize leads)
  9. outreach_list.csv                 (prioritised outreach emails)
  10. unverified_list.csv              (unverified emails)
  11. GS_CRM_WhatsApp_Campaign_Tracker - WhatsApp Tracker sheet
  12. Source CRM Export sheet          - Original CRM contacts

Output:
  GS_Leads_Consolidated_<date>.xlsx  (in same folder as this script)
"""

import pandas as pd
import numpy as np
import re
import warnings
from datetime import datetime
from pathlib import Path

warnings.filterwarnings("ignore")

# ── Paths ────────────────────────────────────────────────────────────────────
DATA_DIR = Path(r"c:\projects\scripts\extract_garden\data")
CRM_DIR  = Path(r"c:\projects\biz\gardensuite.in\marketing\whatsapp-campaigns\face-attendance-weight")
OUT_FILE = CRM_DIR / f"GS_Leads_Consolidated_{datetime.now().strftime('%Y%m%d')}.xlsx"

# ── Helpers ───────────────────────────────────────────────────────────────────
def norm_phone(raw):
    """Return normalised 10-digit Indian mobile as string, or empty string."""
    if pd.isna(raw):
        return ""
    s = re.sub(r"\D", "", str(raw).split(".")[0])
    if s.startswith("91") and len(s) == 12:
        s = s[2:]
    if len(s) == 10 and s[0] in "6789":
        return s
    return ""

def extract_phones(raw):
    """Extract all valid phones from a raw string (for alt numbers fields)."""
    if pd.isna(raw) or str(raw).strip() in ("", "nan"):
        return []
    # pull out digit runs then normalise each
    candidates = re.findall(r"[\d]{7,}", str(raw))
    result = []
    for c in candidates:
        p = norm_phone(c)
        if p and p not in result:
            result.append(p)
    return result

def norm_email(raw):
    if pd.isna(raw) or str(raw).strip() in ("", "nan"):
        return ""
    e = str(raw).strip().lower()
    return e if "@" in e and "." in e.split("@")[1] else ""

def norm_name(raw):
    if pd.isna(raw) or str(raw).strip() in ("", "nan"):
        return ""
    return str(raw).strip()

def norm_estate(raw):
    """Normalised estate key for matching."""
    if not raw:
        return ""
    s = re.sub(r"\s+", " ", str(raw).strip().lower())
    # strip common suffixes for fuzzy match key
    s = re.sub(r"\b(t\.?e\.?|tea estate|pvt\.? ltd\.?|ltd\.?|p\.?v\.?t\.?)\b", "", s)
    return s.strip(" .,")


# ── Union-Find ────────────────────────────────────────────────────────────────
class UnionFind:
    def __init__(self, n):
        self.parent = list(range(n))
        self.rank   = [0] * n

    def find(self, x):
        while self.parent[x] != x:
            self.parent[x] = self.parent[self.parent[x]]
            x = self.parent[x]
        return x

    def union(self, a, b):
        ra, rb = self.find(a), self.find(b)
        if ra == rb:
            return
        if self.rank[ra] < self.rank[rb]:
            ra, rb = rb, ra
        self.parent[rb] = ra
        if self.rank[ra] == self.rank[rb]:
            self.rank[ra] += 1


# ── Load each source ──────────────────────────────────────────────────────────
rows = []   # list of dicts

def add_row(estate, phones, emails, address, website, location, wa_status, crm_status, source, source_rank):
    """Normalise and append a row to the raw list."""
    clean_phones = [p for p in phones if p]
    clean_emails = [e for e in emails if e]
    clean_estate = norm_name(estate)
    # skip entirely empty rows
    if not clean_estate and not clean_phones and not clean_emails:
        return
    rows.append({
        "estate_name":  clean_estate,
        "phones":       clean_phones,   # list
        "emails":       clean_emails,   # list
        "address":      norm_name(address),
        "website":      norm_name(website),
        "location":     norm_name(location),
        "wa_status":    norm_name(wa_status),
        "crm_status":   norm_name(crm_status),
        "source":       source,
        "source_rank":  source_rank,
    })

# --- 1. tea_estate_contacts_v2.xlsx ---
print("Loading tea_estate_contacts_v2.xlsx ...")
df = pd.read_excel(DATA_DIR / "tea_estate_contacts_v2.xlsx")
for _, r in df.iterrows():
    phones = [norm_phone(r.get("GM Phone", ""))]
    phones += extract_phones(r.get("Alternative Numbers", ""))
    phones = list(dict.fromkeys(p for p in phones if p))
    emails = [norm_email(r.get("Emails", ""))]
    emails = [e for e in emails if e]
    add_row(r.get("Tea Estate"), phones, emails,
            r.get("GM Address"), r.get("GM Website"), r.get("Location/District"),
            "", "", "tea_estate_contacts_v2", 1)

# --- 2. tea_estate_contacts.xlsx ---
print("Loading tea_estate_contacts.xlsx ...")
df = pd.read_excel(DATA_DIR / "tea_estate_contacts.xlsx")
for _, r in df.iterrows():
    phones = [norm_phone(r.get("GM Phone", ""))]
    phones += extract_phones(r.get("Alternative Numbers", ""))
    phones = list(dict.fromkeys(p for p in phones if p))
    add_row(r.get("Tea Estate"), phones, [],
            r.get("GM Address"), r.get("GM Website"), "",
            "", "", "tea_estate_contacts", 2)

# --- 3. Tea Estates.xlsx ---
print("Loading Tea Estates.xlsx ...")
df = pd.read_excel(DATA_DIR / "Tea Estates.xlsx", header=1)
name_cols = [c for c in df.columns if "Name" in str(c) and "DPE" in str(c)]
loc_cols  = [c for c in df.columns if "Location" in str(c) or "District" in str(c)]
if name_cols:
    for _, r in df.iterrows():
        name = norm_name(r[name_cols[0]])
        loc  = norm_name(r[loc_cols[0]]) if loc_cols else ""
        if name and name.lower() != "name of dpes (tea estates)":
            add_row(name, [], [], "", "", loc, "", "", "Tea_Estates_list", 5)

# --- 4. Tea Estates number required.xlsx ---
print("Loading Tea Estates number required.xlsx ...")
df = pd.read_excel(DATA_DIR / "Tea Estates number required.xlsx", header=1)
name_cols  = [c for c in df.columns if "Name" in str(c) and "DPE" in str(c)]
phone_cols = [c for c in df.columns if "Unnamed: 5" in str(c)]
if name_cols:
    for _, r in df.iterrows():
        name  = norm_name(r[name_cols[0]])
        phone = norm_phone(r[phone_cols[0]]) if phone_cols else ""
        if name and name.lower() != "name of dpes (tea estates)":
            add_row(name, [phone] if phone else [], [], "", "", "",
                    "", "", "Tea_Estates_number_required", 4)

# --- 5. email assam.dooars teaestate.xlsx ---
print("Loading email assam.dooars teaestate.xlsx ...")
df = pd.read_excel(DATA_DIR / "email assam.dooars teaestate.xlsx")
for _, r in df.iterrows():
    email = norm_email(r.iloc[0])
    if email:
        add_row("", [], [email], "", "", "Assam/Dooars",
                "", "", "email_assam_dooars", 3)

# --- 6. verified_emails_v2.csv ---
print("Loading verified_emails_v2.csv ...")
df = pd.read_csv(DATA_DIR / "verified_emails_v2.csv")
for _, r in df.iterrows():
    email = norm_email(r.get("email", ""))
    if email:
        add_row(r.get("matched_estate"), [], [email], "", "", "",
                "", f"email_verified:{r.get('confidence_level','')}", "verified_emails_v2", 2)

# --- 7. verified_list.csv ---
print("Loading verified_list.csv ...")
df = pd.read_csv(DATA_DIR / "verified_list.csv")
for _, r in df.iterrows():
    email = norm_email(r.get("email", ""))
    if email:
        add_row(r.get("matched_estate"), [], [email], "", "", "",
                "", f"verified:{r.get('confidence_level','')}", "verified_list", 1)

# --- 8. midsize_verified.csv ---
print("Loading midsize_verified.csv ...")
df = pd.read_csv(DATA_DIR / "midsize_verified.csv")
for _, r in df.iterrows():
    email = norm_email(r.get("email", ""))
    if email:
        add_row(r.get("matched_estate"), [], [email], "", "", "",
                "", f"midsize_verified:{r.get('confidence_level','')}", "midsize_verified", 1)

# --- 9. outreach_list.csv ---
print("Loading outreach_list.csv ...")
df = pd.read_csv(DATA_DIR / "outreach_list.csv")
for _, r in df.iterrows():
    email = norm_email(r.get("email", ""))
    if email:
        add_row(r.get("matched_estate"), [], [email], "", "", "",
                "", "outreach_priority", "outreach_list", 1)

# --- 10. unverified_list.csv ---
print("Loading unverified_list.csv ...")
df = pd.read_csv(DATA_DIR / "unverified_list.csv")
for _, r in df.iterrows():
    email = norm_email(r.get("email", ""))
    if email:
        add_row(r.get("matched_estate"), [], [email], "", "", "",
                "", f"unverified:{r.get('confidence_level','')}", "unverified_list", 6)

# --- 11 & 12. GS_CRM_WhatsApp_Campaign_Tracker.xlsx ---
print("Loading GS_CRM_WhatsApp_Campaign_Tracker.xlsx ...")
xl = pd.ExcelFile(CRM_DIR / "GS_CRM_WhatsApp_Campaign_Tracker.xlsx")

df = xl.parse("WhatsApp Tracker", header=0)
for _, r in df.iterrows():
    name  = norm_name(r.get("Client Name")) or norm_name(r.get("Tea Garden"))
    phone = norm_phone(r.get("WhatsApp Number", ""))
    add_row(name, [phone] if phone else [], [], "", "", "",
            norm_name(r.get("Delivery Status")), norm_name(r.get("Original CRM Status")),
            "WA_Campaign_Tracker", 1)

df2 = xl.parse("Source CRM Export", header=0)
for _, r in df2.iterrows():
    name  = norm_name(r.get("Client Name")) or norm_name(r.get("Tea Garden"))
    phone = norm_phone(r.get("Contact", ""))
    add_row(name, [phone] if phone else [], [], "", "", "",
            "", norm_name(r.get("Status")),
            "CRM_Export", 1)


# ── Build raw DataFrame ───────────────────────────────────────────────────────
print(f"\nTotal raw rows collected: {len(rows)}")
all_df = pd.DataFrame(rows)
all_df = all_df.sort_values("source_rank").reset_index(drop=True)
n = len(all_df)


# ── Union-Find clustering ─────────────────────────────────────────────────────
# Two rows are "same target" if they share a phone number, email, OR estate name.
uf = UnionFind(n)

# Build lookup maps: key -> list of row indices
phone_map  = {}
email_map  = {}
estate_map = {}

for i, row in all_df.iterrows():
    estate = norm_estate(row["estate_name"])
    if estate:
        estate_map.setdefault(estate, []).append(i)
    for p in row["phones"]:
        phone_map.setdefault(p, []).append(i)
    for e in row["emails"]:
        email_map.setdefault(e, []).append(i)

# Union all rows that share a phone
for indices in phone_map.values():
    for j in range(1, len(indices)):
        uf.union(indices[0], indices[j])

# Union all rows that share an email
for indices in email_map.values():
    for j in range(1, len(indices)):
        uf.union(indices[0], indices[j])

# Union all rows that share a normalised estate name
for indices in estate_map.values():
    for j in range(1, len(indices)):
        uf.union(indices[0], indices[j])

print(f"Clusters found: {len(set(uf.find(i) for i in range(n)))}")


# ── Merge each cluster into one output row ────────────────────────────────────
from collections import defaultdict

clusters = defaultdict(list)
for i in range(n):
    clusters[uf.find(i)].append(i)

merged = []
for root, members in clusters.items():
    group = all_df.iloc[members].sort_values("source_rank")

    # Collect all unique phones and emails (order preserved, deduped)
    all_phones = []
    all_emails = []
    for _, r in group.iterrows():
        for p in r["phones"]:
            if p and p not in all_phones:
                all_phones.append(p)
        for e in r["emails"]:
            if e and e not in all_emails:
                all_emails.append(e)

    # Best estate name = first non-empty from highest-priority source
    estate_name = next((r["estate_name"] for _, r in group.iterrows() if r["estate_name"]), "")

    # Best address, website, location = first non-empty
    address  = next((r["address"]  for _, r in group.iterrows() if r["address"]),  "")
    website  = next((r["website"]  for _, r in group.iterrows() if r["website"]),  "")
    location = next((r["location"] for _, r in group.iterrows() if r["location"]), "")

    # WA status: keep most informative (non-empty, prefer Sent/Delivered over blank)
    wa_vals = [r["wa_status"] for _, r in group.iterrows() if r["wa_status"]]
    wa_status = wa_vals[0] if wa_vals else ""

    # CRM status: keep most informative
    crm_vals = [r["crm_status"] for _, r in group.iterrows() if r["crm_status"]]
    crm_status = crm_vals[0] if crm_vals else ""

    # Sources (unique, sorted by rank)
    sources = list(dict.fromkeys(r["source"] for _, r in group.iterrows()))

    merged.append({
        "Estate Name": estate_name,
        "Phone 1":    all_phones[0] if len(all_phones) > 0 else "",
        "Phone 2":    all_phones[1] if len(all_phones) > 1 else "",
        "Phone 3":    all_phones[2] if len(all_phones) > 2 else "",
        "Email 1":    all_emails[0] if len(all_emails) > 0 else "",
        "Email 2":    all_emails[1] if len(all_emails) > 1 else "",
        "Email 3":    all_emails[2] if len(all_emails) > 2 else "",
        "Address":    address,
        "Website":    website,
        "Location":   location,
        "WA Status":  wa_status,
        "CRM Status": crm_status,
        "Sources":    ", ".join(sources),
    })

print(f"Unique leads after merging: {len(merged)}")


# ── Final DataFrame ───────────────────────────────────────────────────────────
final = pd.DataFrame(merged)

# Sort: most data first (has phone OR email), then by estate name
final["_score"] = (
    (final["Phone 1"].str.len() == 10).astype(int) * 2 +
    (final["Email 1"].str.contains("@", na=False)).astype(int)
)
final = final.sort_values(["_score", "Estate Name"], ascending=[False, True])
final = final.drop(columns=["_score"]).reset_index(drop=True)
final.index += 1

# Stats
has_p1  = final["Phone 1"].str.len() == 10
has_p2  = final["Phone 2"].str.len() == 10
has_p3  = final["Phone 3"].str.len() == 10
has_e1  = final["Email 1"].str.contains("@", na=False)
has_name = final["Estate Name"].astype(str).str.strip().ne("")
print(f"  Has estate name:    {has_name.sum()}")
print(f"  Has Phone 1:        {has_p1.sum()}")
print(f"  Has Phone 2:        {has_p2.sum()}")
print(f"  Has Phone 3:        {has_p3.sum()}")
print(f"  Has Email 1:        {has_e1.sum()}")
print(f"  Has phone+email:    {(has_p1 & has_e1).sum()}")


# ── Write Excel ───────────────────────────────────────────────────────────────
print(f"\nWriting output to: {OUT_FILE}")
with pd.ExcelWriter(OUT_FILE, engine="openpyxl") as writer:
    final.to_excel(writer, sheet_name="All Leads", index=True, index_label="No.")

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ws = writer.sheets["All Leads"]

    header_fill  = PatternFill("solid", fgColor="1B4332")  # dark green
    header_font  = Font(color="FFFFFF", bold=True, size=11)
    alt_fill     = PatternFill("solid", fgColor="F0F4F1")
    phone_fill   = PatternFill("solid", fgColor="E8F5E9")  # light green tint for phone cols
    email_fill   = PatternFill("solid", fgColor="E3F2FD")  # light blue tint for email cols
    border_side  = Side(style="thin", color="CCCCCC")
    thin_border  = Border(bottom=border_side)

    col_widths = {
        "A":  6,   # No.
        "B": 28,   # Estate Name
        "C": 14,   # Phone 1
        "D": 14,   # Phone 2
        "E": 14,   # Phone 3
        "F": 34,   # Email 1
        "G": 34,   # Email 2
        "H": 34,   # Email 3
        "I": 40,   # Address
        "J": 22,   # Website
        "K": 18,   # Location
        "L": 14,   # WA Status
        "M": 22,   # CRM Status
        "N": 36,   # Sources
    }

    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Phone cols: C, D, E  (index 3,4,5)  Email cols: F, G, H (index 6,7,8)
    PHONE_COLS = {3, 4, 5}
    EMAIL_COLS = {6, 7, 8}

    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        for col_idx, cell in enumerate(row, start=1):
            if row_idx == 1:
                cell.font      = header_font
                cell.fill      = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center", wrap_text=False)
                if col_idx in PHONE_COLS:
                    cell.fill = phone_fill
                elif col_idx in EMAIL_COLS:
                    cell.fill = email_fill
                elif row_idx % 2 == 0:
                    cell.fill = alt_fill
            cell.border = thin_border

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 22

print("Done!")
print(f"  Output : {OUT_FILE}")
print(f"  Rows   : {len(final)}")
