"""
GardenSuite Lead Consolidation Script
======================================
Merges all lead sources into one clean Excel file with duplicates removed.

Sources:
  1. tea_estate_contacts_v2.xlsx       - 222 rows (best: phone + email + address)
  2. tea_estate_contacts.xlsx          - 173 rows (phone + address)
  3. Tea Estates.xlsx                  - 174 rows (names + area)
  4. Tea Estates number required.xlsx  - 84 rows  (names + phone)
  5. email assam.dooars teaestate.xlsx - 149 rows (emails only)
  6. verified_emails_v2.csv            - 145 rows (verified emails + estate match)
  7. verified_list.csv                 - 31 rows  (high-confidence verified emails)
  8. midsize_verified.csv              - 19 rows  (verified midsize leads)
  9. outreach_list.csv                 - 17 rows  (prioritised outreach emails)
  10. unverified_list.csv              - 114 rows (unverified emails)
  11. GS_CRM_WhatsApp_Campaign_Tracker - WA campaign contacts (WhatsApp Tracker sheet)
  12. Source CRM Export sheet          - Original CRM contacts

Output:
  GS_Leads_Consolidated_<date>.xlsx  (in the whatsapp-campaigns folder)

Deduplication logic (priority order):
  1. Deduplicate by phone (normalised 10-digit)
  2. Deduplicate by email (lowercased)
  3. Deduplicate by estate name (lowercased, stripped)
  Higher-priority sources win (lower source_rank = higher priority).
"""

import pandas as pd
import numpy as np
import re
import warnings
from datetime import datetime
from pathlib import Path

warnings.filterwarnings("ignore")

# ── Paths ────────────────────────────────────────────────────────────────────
DATA_DIR = Path(r"c:\projects\scripts\extract_garden\data")
CRM_DIR  = Path(r"c:\projects\biz\gardensuite.in\marketing\whatsapp-campaigns\face-attendance-weight")
OUT_FILE = CRM_DIR / f"GS_Leads_Consolidated_{datetime.now().strftime('%Y%m%d')}.xlsx"

# ── Helpers ───────────────────────────────────────────────────────────────────
def norm_phone(raw):
    """Return normalised 10-digit Indian mobile as string, or empty string."""
    if pd.isna(raw):
        return ""
    s = re.sub(r"\D", "", str(raw).split(".")[0])  # strip float decimal
    if s.startswith("91") and len(s) == 12:
        s = s[2:]
    if len(s) == 10 and s[0] in "6789":
        return s
    return ""

def norm_email(raw):
    if pd.isna(raw) or str(raw).strip() in ("", "nan"):
        return ""
    e = str(raw).strip().lower()
    return e if "@" in e else ""

def norm_name(raw):
    if pd.isna(raw) or str(raw).strip() in ("", "nan"):
        return ""
    return str(raw).strip()

def norm_estate(raw):
    if not raw:
        return ""
    return re.sub(r"\s+", " ", str(raw).strip().lower())


# ── Load each source ──────────────────────────────────────────────────────────
rows = []   # list of dicts

# --- 1. tea_estate_contacts_v2.xlsx (best source) ---
print("Loading tea_estate_contacts_v2.xlsx ...")
df = pd.read_excel(DATA_DIR / "tea_estate_contacts_v2.xlsx")
for _, r in df.iterrows():
    phone = norm_phone(r.get("GM Phone", ""))
    alt_raw = str(r.get("Alternative Numbers", "")) if not pd.isna(r.get("Alternative Numbers", np.nan)) else ""
    extra_phones = [norm_phone(p) for p in re.findall(r"[\d]{7,}", alt_raw)]
    email = norm_email(r.get("Emails", ""))
    row = {
        "estate_name":  norm_name(r.get("Tea Estate", "")),
        "phone":        phone,
        "alt_phones":   "; ".join(p for p in extra_phones if p),
        "email":        email,
        "address":      norm_name(r.get("GM Address", "")),
        "website":      norm_name(r.get("GM Website", "")),
        "location":     norm_name(r.get("Location/District", "")),
        "wa_status":    "",
        "crm_status":   "",
        "source":       "tea_estate_contacts_v2",
        "source_rank":  1,
    }
    if row["estate_name"] or row["phone"] or row["email"]:
        rows.append(row)

# --- 2. tea_estate_contacts.xlsx ---
print("Loading tea_estate_contacts.xlsx ...")
df = pd.read_excel(DATA_DIR / "tea_estate_contacts.xlsx")
for _, r in df.iterrows():
    estate = norm_name(r.get("Tea Estate", ""))
    phone  = norm_phone(r.get("GM Phone", ""))
    if estate or phone:
        rows.append({
            "estate_name":  estate,
            "phone":        phone,
            "alt_phones":   "",
            "email":        "",
            "address":      norm_name(r.get("GM Address", "")),
            "website":      norm_name(r.get("GM Website", "")),
            "location":     "",
            "wa_status":    "",
            "crm_status":   "",
            "source":       "tea_estate_contacts",
            "source_rank":  2,
        })

# --- 3. Tea Estates.xlsx (skip 2 header rows) ---
print("Loading Tea Estates.xlsx ...")
df = pd.read_excel(DATA_DIR / "Tea Estates.xlsx", header=1)
for _, r in df.iterrows():
    # column is named "Name of DPEs (Tea Estates)" after header=1
    name_cols = [c for c in df.columns if "Name" in str(c) and "DPE" in str(c)]
    loc_cols  = [c for c in df.columns if "Location" in str(c) or "District" in str(c)]
    name = norm_name(r[name_cols[0]]) if name_cols else ""
    loc  = norm_name(r[loc_cols[0]])  if loc_cols  else ""
    if name and name.lower() not in ("name of dpes (tea estates)",):
        rows.append({
            "estate_name":  name,
            "phone":        "",
            "alt_phones":   "",
            "email":        "",
            "address":      "",
            "website":      "",
            "location":     loc,
            "wa_status":    "",
            "crm_status":   "",
            "source":       "Tea_Estates_list",
            "source_rank":  5,
        })

# --- 4. Tea Estates number required.xlsx ---
print("Loading Tea Estates number required.xlsx ...")
df = pd.read_excel(DATA_DIR / "Tea Estates number required.xlsx", header=1)
name_cols  = [c for c in df.columns if "Name" in str(c) and "DPE" in str(c)]
phone_cols = [c for c in df.columns if "Unnamed: 5" in str(c) or "Phone" in str(c) or "Number" in str(c)]
for _, r in df.iterrows():
    name  = norm_name(r[name_cols[0]])  if name_cols  else ""
    phone = norm_phone(r[phone_cols[0]]) if phone_cols else ""
    if name and name.lower() not in ("name of dpes (tea estates)",):
        rows.append({
            "estate_name":  name,
            "phone":        phone,
            "alt_phones":   "",
            "email":        "",
            "address":      "",
            "website":      "",
            "location":     "",
            "wa_status":    "",
            "crm_status":   "",
            "source":       "Tea_Estates_number_required",
            "source_rank":  4,
        })

# --- 5. email assam.dooars teaestate.xlsx (emails only) ---
print("Loading email assam.dooars teaestate.xlsx ...")
df = pd.read_excel(DATA_DIR / "email assam.dooars teaestate.xlsx")
email_col = df.columns[0]
for _, r in df.iterrows():
    email = norm_email(r[email_col])
    if email:
        rows.append({
            "estate_name":  "",
            "phone":        "",
            "alt_phones":   "",
            "email":        email,
            "address":      "",
            "website":      "",
            "location":     "Assam/Dooars",
            "wa_status":    "",
            "crm_status":   "",
            "source":       "email_assam_dooars",
            "source_rank":  3,
        })

# --- 6. verified_emails_v2.csv ---
print("Loading verified_emails_v2.csv ...")
df = pd.read_csv(DATA_DIR / "verified_emails_v2.csv")
for _, r in df.iterrows():
    email = norm_email(r.get("email", ""))
    if email:
        rows.append({
            "estate_name":  norm_name(r.get("matched_estate", "")),
            "phone":        "",
            "alt_phones":   "",
            "email":        email,
            "address":      "",
            "website":      "",
            "location":     "",
            "wa_status":    "",
            "crm_status":   f"email_verified:{r.get('confidence_level','')}",
            "source":       "verified_emails_v2",
            "source_rank":  2,
        })

# --- 7. verified_list.csv ---
print("Loading verified_list.csv ...")
df = pd.read_csv(DATA_DIR / "verified_list.csv")
for _, r in df.iterrows():
    email = norm_email(r.get("email", ""))
    if email:
        rows.append({
            "estate_name":  norm_name(r.get("matched_estate", "")),
            "phone":        "",
            "alt_phones":   "",
            "email":        email,
            "address":      "",
            "website":      "",
            "location":     "",
            "wa_status":    "",
            "crm_status":   f"verified:{r.get('confidence_level','')}",
            "source":       "verified_list",
            "source_rank":  1,
        })

# --- 8. midsize_verified.csv ---
print("Loading midsize_verified.csv ...")
df = pd.read_csv(DATA_DIR / "midsize_verified.csv")
for _, r in df.iterrows():
    email = norm_email(r.get("email", ""))
    if email:
        rows.append({
            "estate_name":  norm_name(r.get("matched_estate", "")),
            "phone":        "",
            "alt_phones":   "",
            "email":        email,
            "address":      "",
            "website":      "",
            "location":     "",
            "wa_status":    "",
            "crm_status":   f"midsize_verified:{r.get('confidence_level','')}",
            "source":       "midsize_verified",
            "source_rank":  1,
        })

# --- 9. outreach_list.csv ---
print("Loading outreach_list.csv ...")
df = pd.read_csv(DATA_DIR / "outreach_list.csv")
for _, r in df.iterrows():
    email = norm_email(r.get("email", ""))
    if email:
        rows.append({
            "estate_name":  norm_name(r.get("matched_estate", "")),
            "phone":        "",
            "alt_phones":   "",
            "email":        email,
            "address":      "",
            "website":      "",
            "location":     "",
            "wa_status":    "",
            "crm_status":   "outreach_priority",
            "source":       "outreach_list",
            "source_rank":  1,
        })

# --- 10. unverified_list.csv ---
print("Loading unverified_list.csv ...")
df = pd.read_csv(DATA_DIR / "unverified_list.csv")
for _, r in df.iterrows():
    email = norm_email(r.get("email", ""))
    if email:
        rows.append({
            "estate_name":  norm_name(r.get("matched_estate", "")),
            "phone":        "",
            "alt_phones":   "",
            "email":        email,
            "address":      "",
            "website":      "",
            "location":     "",
            "wa_status":    "",
            "crm_status":   f"unverified:{r.get('confidence_level','')}",
            "source":       "unverified_list",
            "source_rank":  6,
        })

# --- 11. GS_CRM_WhatsApp_Campaign_Tracker - WhatsApp Tracker sheet ---
print("Loading GS_CRM_WhatsApp_Campaign_Tracker.xlsx ...")
xl = pd.ExcelFile(CRM_DIR / "GS_CRM_WhatsApp_Campaign_Tracker.xlsx")

df = xl.parse("WhatsApp Tracker", header=0)
for _, r in df.iterrows():
    name  = norm_name(r.get("Client Name", "")) or norm_name(r.get("Tea Garden", ""))
    phone = norm_phone(r.get("WhatsApp Number", ""))
    if name or phone:
        rows.append({
            "estate_name":  name,
            "phone":        phone,
            "alt_phones":   "",
            "email":        "",
            "address":      "",
            "website":      "",
            "location":     "",
            "wa_status":    norm_name(r.get("Delivery Status", "")),
            "crm_status":   norm_name(r.get("Original CRM Status", "")),
            "source":       "WA_Campaign_Tracker",
            "source_rank":  1,
        })

# --- 12. Source CRM Export sheet ---
df2 = xl.parse("Source CRM Export", header=0)
for _, r in df2.iterrows():
    name  = norm_name(r.get("Client Name", "")) or norm_name(r.get("Tea Garden", ""))
    phone = norm_phone(r.get("Contact", ""))
    if name or phone:
        rows.append({
            "estate_name":  name,
            "phone":        phone,
            "alt_phones":   "",
            "email":        "",
            "address":      "",
            "website":      "",
            "location":     "",
            "wa_status":    "",
            "crm_status":   norm_name(r.get("Status", "")),
            "source":       "CRM_Export",
            "source_rank":  1,
        })


# ── Build DataFrame ───────────────────────────────────────────────────────────
print(f"\nTotal raw rows collected: {len(rows)}")
all_df = pd.DataFrame(rows)

# Sort by source_rank ascending (lower = higher priority wins during dedup)
all_df = all_df.sort_values("source_rank").reset_index(drop=True)

# ── Deduplication ─────────────────────────────────────────────────────────────
# For each row, check if we've already seen this phone / email / estate.
# If yes, merge any missing fields into the existing winner and skip this row.
# If no, add as a new winner.

seen_phones  = {}  # phone -> index in all_df
seen_emails  = {}  # email -> index in all_df
seen_estates = {}  # normalised estate -> index in all_df
keep_mask    = [False] * len(all_df)

for i in range(len(all_df)):
    r      = all_df.iloc[i]
    phone  = r["phone"]
    email  = r["email"]
    estate = norm_estate(r["estate_name"])

    winner_i = None

    if phone and phone in seen_phones:
        winner_i = seen_phones[phone]
    elif email and email in seen_emails:
        winner_i = seen_emails[email]
    elif estate and estate in seen_estates:
        winner_i = seen_estates[estate]

    if winner_i is not None:
        # Merge missing fields into the winner
        for field in ["estate_name", "email", "phone", "address", "location", "website", "wa_status", "crm_status"]:
            if not all_df.at[winner_i, field] and r[field]:
                all_df.at[winner_i, field] = r[field]
        if not all_df.at[winner_i, "alt_phones"] and r["alt_phones"]:
            all_df.at[winner_i, "alt_phones"] = r["alt_phones"]
    else:
        keep_mask[i] = True
        if phone:  seen_phones[phone]   = i
        if email:  seen_emails[email]   = i
        if estate: seen_estates[estate] = i

deduped = all_df[keep_mask].copy()
print(f"After deduplication: {len(deduped)} unique leads")

# ── Final clean-up ────────────────────────────────────────────────────────────
final = deduped[[
    "estate_name", "phone", "alt_phones", "email",
    "address", "website", "location",
    "wa_status", "crm_status", "source"
]].copy()

final.columns = [
    "Estate Name", "Phone", "Alt Phones", "Email",
    "Address", "Website", "Location",
    "WA Status", "CRM Status", "Source"
]

# Sort: contacts with phone or email first, then alphabetically by estate name
final["_score"] = (
    (final["Phone"].astype(str).str.len() == 10).astype(int) * 2 +
    (final["Email"].str.contains("@", na=False)).astype(int)
)
final = final.sort_values(["_score", "Estate Name"], ascending=[False, True])
final = final.drop(columns=["_score"]).reset_index(drop=True)
final.index += 1

# ── Stats ─────────────────────────────────────────────────────────────────────
has_phone  = final["Phone"].str.len() == 10
has_email  = final["Email"].str.contains("@", na=False)
has_name   = final["Estate Name"].astype(str).str.strip().ne("")
print(f"  Has phone:       {has_phone.sum()}")
print(f"  Has email:       {has_email.sum()}")
print(f"  Has estate name: {has_name.sum()}")
print(f"  Has phone+email: {(has_phone & has_email).sum()}")
print(f"  Source breakdown:")
print(final.groupby("Source").size().sort_values(ascending=False).to_string())

# ── Write Excel ───────────────────────────────────────────────────────────────
print(f"\nWriting output to: {OUT_FILE}")
with pd.ExcelWriter(OUT_FILE, engine="openpyxl") as writer:
    final.to_excel(writer, sheet_name="All Leads", index=True, index_label="No.")

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = writer.book
    ws = writer.sheets["All Leads"]

    header_fill = PatternFill("solid", fgColor="1B4332")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    alt_fill    = PatternFill("solid", fgColor="F0F4F1")
    border_side = Side(style="thin", color="CCCCCC")
    thin_border = Border(bottom=border_side)

    col_widths = {
        "A": 6,   # No.
        "B": 28,  # Estate Name
        "C": 14,  # Phone
        "D": 28,  # Alt Phones
        "E": 34,  # Email
        "F": 40,  # Address
        "G": 22,  # Website
        "H": 18,  # Location
        "I": 14,  # WA Status
        "J": 22,  # CRM Status
        "K": 26,  # Source
    }

    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        for cell in row:
            if row_idx == 1:
                cell.font      = header_font
                cell.fill      = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center", wrap_text=False)
                if row_idx % 2 == 0:
                    cell.fill = alt_fill
            cell.border = thin_border

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 22

print("Done!")
print(f"  Output : {OUT_FILE}")
print(f"  Rows   : {len(final)}")
