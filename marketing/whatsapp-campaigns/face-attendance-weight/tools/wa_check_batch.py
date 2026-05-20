import argparse
import csv
import random
import re
import sys
import time
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
from playwright.sync_api import sync_playwright


ROOT = Path(__file__).resolve().parents[4]
CAMPAIGN_DIR = ROOT / "marketing" / "whatsapp-campaigns" / "face-attendance-weight"
WORKBOOK = CAMPAIGN_DIR / "GS_CRM_WhatsApp_Campaign_Tracker.xlsx"
RESULTS_DIR = CAMPAIGN_DIR / "wa-check-results"
PROFILE_DIR = CAMPAIGN_DIR / ".wa-check-browser-profile"

NO_WA_PATTERNS = [
    "phone number shared via url is invalid",
    "phone number shared via url is invalid.",
    "invalid phone number",
    "not on whatsapp",
    "number is not on whatsapp",
]
LOGIN_PATTERNS = [
    "use whatsapp on your computer",
    "log in to whatsapp",
    "link a device",
    "scan the qr code",
]
CHAT_SELECTORS = [
    'div[contenteditable="true"][data-tab]',
    'footer div[contenteditable="true"]',
    '[aria-label="Type a message"]',
    '[aria-placeholder="Type a message"]',
]


def normalize_phone(value):
    if value is None:
        return ""
    digits = re.sub(r"\D", "", str(value))
    if len(digits) == 12 and digits.startswith("91"):
        digits = digits[2:]
    if len(digits) == 11 and digits.startswith("0"):
        digits = digits[1:]
    if len(digits) == 10 and digits[0] in "6789":
        return digits
    return ""


def get_headers(ws):
    return {str(ws.cell(1, col).value).strip(): col for col in range(1, ws.max_column + 1) if ws.cell(1, col).value}


def select_batch(limit, include_hold=False):
    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    ws = wb["WhatsApp Tracker"]
    h = get_headers(ws)
    required = ["Lead ID", "Client Name", "Tea Garden", "WhatsApp Number", "Send Decision", "WA Check Link", "WA Check Status"]
    missing = [name for name in required if name not in h]
    if missing:
        raise SystemExit(f"Workbook is missing columns: {', '.join(missing)}")

    rows = []
    for row in range(2, ws.max_row + 1):
        status = str(ws.cell(row, h["WA Check Status"]).value or "Unchecked").strip()
        send_decision = str(ws.cell(row, h["Send Decision"]).value or "").strip()
        if status not in {"", "Unchecked", "Check failed", "Login required"}:
            continue
        if not include_hold and send_decision == "Hold":
            continue
        phone = normalize_phone(ws.cell(row, h["WhatsApp Number"]).value)
        if not phone:
            continue
        rows.append(
            {
                "row": row,
                "lead_id": ws.cell(row, h["Lead ID"]).value or "",
                "client_name": ws.cell(row, h["Client Name"]).value or "",
                "tea_garden": ws.cell(row, h["Tea Garden"]).value or "",
                "phone": phone,
                "wa_link": f"https://wa.me/91{phone}",
                "send_decision": send_decision,
            }
        )
        if len(rows) >= limit:
            break
    return rows


def classify_page(page):
    text = ""
    try:
        text = page.locator("body").inner_text(timeout=5000)
    except Exception:
        pass
    lower = text.lower()
    if any(pattern in lower for pattern in NO_WA_PATTERNS):
        return "No WhatsApp", "WhatsApp Web showed invalid/not-on-WhatsApp message"
    if any(pattern in lower for pattern in LOGIN_PATTERNS):
        return "Login required", "WhatsApp Web login required, scan QR and rerun"
    for selector in CHAT_SELECTORS:
        try:
            if page.locator(selector).first.is_visible(timeout=1500):
                return "Has WhatsApp", "Chat input detected"
        except Exception:
            continue
    if "web.whatsapp.com/send" in page.url.lower() and "phone=" in page.url.lower():
        return "Check failed", "Reached WhatsApp Web but could not confirm chat or invalid state"
    return "Check failed", "Could not classify page"


def run_check(rows, delay_min, delay_max, headless=False, dry_run=False):
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    out = RESULTS_DIR / f"wa_check_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

    fields = [
        "row",
        "lead_id",
        "client_name",
        "tea_garden",
        "phone",
        "wa_link",
        "send_decision",
        "wa_check_status",
        "wa_checked_at",
        "wa_check_notes",
    ]

    if dry_run:
        with out.open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fields)
            writer.writeheader()
            for item in rows:
                writer.writerow({**item, "wa_check_status": "Dry run", "wa_checked_at": "", "wa_check_notes": "Not checked"})
        return out

    with sync_playwright() as p:
        context = p.chromium.launch_persistent_context(
            str(PROFILE_DIR),
            headless=headless,
            viewport={"width": 1280, "height": 900},
            args=["--disable-blink-features=AutomationControlled"],
        )
        page = context.new_page()
        with out.open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fields)
            writer.writeheader()
            for index, item in enumerate(rows, start=1):
                status = "Check failed"
                notes = ""
                try:
                    # First load the requested wa.me URL. If WhatsApp presents an intermediate
                    # page, try the web endpoint after a short wait. No message is sent.
                    page.goto(item["wa_link"], wait_until="domcontentloaded", timeout=30000)
                    time.sleep(2)
                    if "wa.me" in page.url.lower() or "api.whatsapp.com" in page.url.lower():
                        page.goto(
                            f"https://web.whatsapp.com/send?phone=91{item['phone']}&text&app_absent=0",
                            wait_until="domcontentloaded",
                            timeout=45000,
                        )
                    try:
                        page.wait_for_load_state("networkidle", timeout=15000)
                    except PlaywrightTimeoutError:
                        pass
                    time.sleep(6)
                    status, notes = classify_page(page)
                except Exception as exc:
                    status = "Check failed"
                    notes = f"{type(exc).__name__}: {exc}"

                writer.writerow(
                    {
                        **item,
                        "wa_check_status": status,
                        "wa_checked_at": datetime.now().isoformat(timespec="seconds"),
                        "wa_check_notes": notes,
                    }
                )
                f.flush()
                if index < len(rows):
                    time.sleep(random.uniform(delay_min, delay_max))
        context.close()
    return out


def main():
    parser = argparse.ArgumentParser(description="Safely check a small batch of WhatsApp numbers into a review CSV.")
    parser.add_argument("--limit", type=int, default=25, help="Batch size. Hard capped at 50.")
    parser.add_argument("--delay-min", type=float, default=8, help="Minimum delay between checks.")
    parser.add_argument("--delay-max", type=float, default=15, help="Maximum delay between checks.")
    parser.add_argument("--include-hold", action="store_true", help="Also check rows where Send Decision is Hold.")
    parser.add_argument("--headless", action="store_true", help="Run browser hidden. Not recommended for first login.")
    parser.add_argument("--dry-run", action="store_true", help="Only export selected rows. Do not open browser.")
    args = parser.parse_args()

    if args.limit > 50:
        raise SystemExit("Safety stop: limit is capped at 50 per run.")
    if args.limit < 1:
        raise SystemExit("Limit must be at least 1.")
    if args.delay_min < 8:
        raise SystemExit("Safety stop: delay-min must be at least 8 seconds.")
    if args.delay_max < args.delay_min:
        raise SystemExit("delay-max must be greater than or equal to delay-min.")

    rows = select_batch(args.limit, include_hold=args.include_hold)
    if not rows:
        print("No rows selected for checking.")
        return

    out = run_check(rows, args.delay_min, args.delay_max, headless=args.headless, dry_run=args.dry_run)
    print(f"Selected rows: {len(rows)}")
    print(f"Results CSV: {out}")
    if not args.dry_run:
        print("Review the CSV, then run apply-wa-check-results.ps1 with that file.")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"ERROR: {type(exc).__name__}: {exc}", file=sys.stderr)
        raise
